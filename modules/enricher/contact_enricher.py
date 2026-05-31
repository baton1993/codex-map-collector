#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Contact cleaner and website enrichment engine for exported company tables."""

from __future__ import annotations

import argparse
import asyncio
import csv
import html
import io
import json
import re
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import unquote, urljoin, urlsplit, urlunsplit

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
INPUT_DIR = ROOT / "input"
OUTPUT_DIR = ROOT / "output"
SUPPORTED_EXTENSIONS = {".xlsx", ".csv", ".json", ".docx"}

USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36"
)

OUTPUT_COLUMNS = [
    "Название",
    "Сайт",
    "Соцсети",
    "Ссылка 2ГИС",
    "Почта электронная",
    "Номер",
    "Статус сайта",
    "HTTP код",
    "Свежесть балл",
    "Год в футере",
    "Сегмент",
    "Источник сайта",
    "Проверено",
    "Комментарий",
]

FIELD_ALIASES = {
    "name": ("название", "компания", "организация", "name", "company"),
    "website": ("сайт", "website", "site", "web"),
    "email": ("email", "e-mail", "почта", "электроннаяпочта"),
    "socials": ("соцсети", "соцсеть", "social", "socials", "links"),
    "phone": ("телефоны", "телефон", "номер", "phone", "phones"),
    "gis_link": ("2gisссылка", "2гисссылка", "2gis", "2гис", "дубльгис"),
    "address": ("адрес", "address"),
    "categories": ("рубрики", "категории", "category", "categories"),
    "rating": ("рейтинг", "rating"),
    "reviews": ("колвоотзывов", "отзывы", "reviews"),
}

SOCIAL_HOSTS = {
    "vk.com": "ВК",
    "m.vk.com": "ВК",
    "t.me": "Telegram",
    "telegram.me": "Telegram",
    "wa.me": "WhatsApp",
    "api.whatsapp.com": "WhatsApp",
    "whatsapp.com": "WhatsApp",
    "mssg.me": "WhatsApp",
    "me-qr.com": "WhatsApp",
    "instagram.com": "Instagram",
    "ok.ru": "ОК",
    "youtube.com": "YouTube",
    "youtu.be": "YouTube",
    "dzen.ru": "Дзен",
    "facebook.com": "Facebook",
    "rutube.ru": "Rutube",
    "tenchat.ru": "TenChat",
    "viber.com": "Viber",
    "twitter.com": "Twitter",
    "x.com": "Twitter",
    "taplink.cc": "Taplink",
    "taplink.ru": "Taplink",
    "taplink.me": "Taplink",
    "linktr.ee": "Taplink",
    "linkinbio.com": "Taplink",
    "beacons.ai": "Taplink",
    "max.ru": "Max",
    "max.com": "Max",
}

NOISE_SITE_HOSTS = (
    "2gis.ru",
    "link.2gis.ru",
    "google.com",
    "yandex.ru",
    "yandex.com",
    "maps.google.com",
)

PARKING_HOST_PARTS = (
    "sedoparking",
    "parkingcrew",
    "bodis.com",
    "above.com",
    "dan.com",
    "afternic",
    "hugedomains",
    "reg.ru/domain",
)

PARKING_WORDS = (
    "domain is for sale",
    "this domain is for sale",
    "buy this domain",
    "домен продается",
    "домен продаётся",
    "купить этот домен",
    "купить домен",
    "домен припаркован",
    "domain parking",
    "account suspended",
    "сайт приостановлен",
    "хостинг приостановлен",
)

PROTECTION_WORDS = (
    "cloudflare",
    "ddos-guard",
    "captcha",
    "recaptcha",
    "checking your browser",
    "just a moment",
    "enable cookies",
    "доступ ограничен",
    "проверка браузера",
    "защита от ботов",
)

CONTACT_PATHS = (
    "/contacts",
    "/contact",
    "/kontakty",
    "/kontakti",
    "/контакты",
    "/about",
)

CONTACT_WORDS = (
    "contact",
    "contacts",
    "kontakty",
    "kontakti",
    "контакт",
    "связ",
    "реквизит",
)

EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
URL_RE = re.compile(r"(?i)\b(?:https?://|www\.)[^\s<>'\"\]\[{};]+")
DOMAIN_RE = re.compile(
    r"(?i)(?<![@/\w.-])((?:[a-z0-9-]+\.)+"
    r"(?:ru|рф|com|net|org|su|biz|info|online|site|shop|pro)"
    r"(?:/[^\s;]*)?)"
)

EMAIL_NOISE = (
    ".png",
    ".jpg",
    ".jpeg",
    ".gif",
    ".webp",
    ".svg",
    "example.com",
    "domain.com",
    "yourdomain",
    "sentry",
    "wixpress",
    "godaddy",
    "u003e",
    "@2x",
    "@3x",
)

TECHNICAL_EMAIL_DOMAINS = {
    "beget.com",
    "constructor.ru",
}

GOOD_SITE_STATUSES = {
    "живой",
    "открылся частично",
    "доступ ограничен/защита",
    "ошибка сервера",
}


@dataclass
class Record:
    row_number: int
    name: str
    raw_website: str
    raw_email: str
    raw_socials: str
    raw_phone: str
    gis_link: str
    address: str = ""
    categories: str = ""
    rating: str = ""
    reviews: str = ""


def clean_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).replace("\xa0", " ").strip()
    return "" if text.lower() in {"nan", "none"} else text


def normalize_header(value: Any) -> str:
    text = clean_cell(value).lower().replace("ё", "е")
    return re.sub(r"[^a-zа-я0-9]+", "", text)


def unique(items: list[str] | tuple[str, ...] | set[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for item in items:
        if not item:
            continue
        key = item.strip().lower()
        if not key or key in seen:
            continue
        seen.add(key)
        out.append(item.strip())
    return out


def detect_columns_from_rows(rows: list[list[Any]]) -> tuple[int, dict[str, int]]:
    best_row = 0
    best_score = -1
    best_map: dict[str, int] = {}

    alias_norm = {
        field: tuple(normalize_header(alias) for alias in aliases)
        for field, aliases in FIELD_ALIASES.items()
    }

    for row_index, values in enumerate(rows[:10]):
        current: dict[str, int] = {}
        for idx, value in enumerate(values):
            key = normalize_header(value)
            if not key:
                continue
            for field, aliases in alias_norm.items():
                exact = key in aliases
                contains = any(alias and alias in key for alias in aliases)
                if (exact or contains) and field not in current:
                    current[field] = idx
        score = len(current)
        if score > best_score:
            best_score = score
            best_row = row_index
            best_map = current

    if "name" not in best_map:
        raise ValueError("Не нашёл колонку с названием компании.")
    return best_row, best_map


def records_from_rows(rows: list[list[Any]]) -> list[Record]:
    header_row, columns = detect_columns_from_rows(rows)
    records: list[Record] = []

    def get(row: list[Any], field: str) -> str:
        idx = columns.get(field)
        if idx is None or idx >= len(row):
            return ""
        return clean_cell(row[idx])

    for row_number, row in enumerate(rows[header_row + 1 :], start=header_row + 2):
        if not any(clean_cell(value) for value in row):
            continue
        name = get(row, "name")
        if not name:
            continue
        records.append(
            Record(
                row_number=row_number,
                name=name,
                raw_website=get(row, "website"),
                raw_email=get(row, "email"),
                raw_socials=get(row, "socials"),
                raw_phone=get(row, "phone"),
                gis_link=get(row, "gis_link"),
                address=get(row, "address"),
                categories=get(row, "categories"),
                rating=get(row, "rating"),
                reviews=get(row, "reviews"),
            )
        )
    return records


def read_xlsx_records(path: Path, sheet_name: str | None = None) -> list[Record]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    return records_from_rows(rows)


def read_text_with_fallback(path: Path) -> str:
    errors: list[str] = []
    for encoding in ("utf-8-sig", "utf-8", "cp1251", "windows-1251"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError as exc:
            errors.append(f"{encoding}: {exc}")
    raise UnicodeDecodeError(
        "utf-8/cp1251",
        b"",
        0,
        1,
        "не смог прочитать файл как UTF-8 или Windows-1251",
    )


def read_csv_records(path: Path) -> list[Record]:
    text = read_text_with_fallback(path)
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        rows = [row for row in csv.reader(io.StringIO(text), dialect) if any(cell.strip() for cell in row)]
    except csv.Error:
        rows = [row for row in csv.reader(io.StringIO(text), delimiter=";") if any(cell.strip() for cell in row)]
    return records_from_rows(rows)


def json_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return clean_cell(value)
    if isinstance(value, list) and all(not isinstance(item, (dict, list)) for item in value):
        return "; ".join(clean_cell(item) for item in value if clean_cell(item))
    return json.dumps(value, ensure_ascii=False)


def flatten_json_object(obj: dict[str, Any], prefix: str = "") -> dict[str, str]:
    out: dict[str, str] = {}
    for key, value in obj.items():
        clean_key = clean_cell(key)
        full_key = f"{prefix}.{clean_key}" if prefix else clean_key
        if isinstance(value, dict):
            out.update(flatten_json_object(value, full_key))
        else:
            out[full_key] = json_cell(value)
    return out


def find_json_records(data: Any) -> list[dict[str, Any]]:
    if isinstance(data, list):
        return [item for item in data if isinstance(item, dict)]
    if isinstance(data, dict):
        name_aliases = {normalize_header(alias) for alias in FIELD_ALIASES["name"]}
        if any(normalize_header(key) in name_aliases for key in data):
            return [data]
        for key in ("items", "data", "rows", "records", "companies", "results"):
            value = data.get(key)
            if isinstance(value, list) and any(isinstance(item, dict) for item in value):
                return [item for item in value if isinstance(item, dict)]
        best: list[dict[str, Any]] = []
        for value in data.values():
            nested = find_json_records(value)
            if len(nested) > len(best):
                best = nested
        return best
    return []


def read_json_records(path: Path) -> list[Record]:
    data = json.loads(read_text_with_fallback(path))
    items = find_json_records(data)
    if not items:
        raise ValueError("В JSON не нашёл список компаний.")
    flattened = [flatten_json_object(item) for item in items]
    headers: list[str] = []
    for item in flattened:
        headers.extend(key for key in item if key not in headers)
    rows = [headers] + [[item.get(header, "") for header in headers] for item in flattened]
    return records_from_rows(rows)


def read_docx_records(path: Path) -> list[Record]:
    try:
        from docx import Document
    except Exception as exc:
        raise RuntimeError(
            "Для чтения DOCX нужен пакет python-docx. Запусти run.command, он установит зависимости."
        ) from exc

    document = Document(path)
    candidates: list[list[Record]] = []
    for table in document.tables:
        rows = [[cell.text.strip() for cell in row.cells] for row in table.rows]
        if not rows:
            continue
        try:
            records = records_from_rows(rows)
        except ValueError:
            continue
        if records:
            candidates.append(records)

    if candidates:
        return max(candidates, key=len)

    raise ValueError("В DOCX не нашёл таблицу с колонкой названия компании.")


def read_records(path: Path, sheet_name: str | None = None) -> list[Record]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return read_xlsx_records(path, sheet_name)
    if suffix == ".csv":
        return read_csv_records(path)
    if suffix == ".json":
        return read_json_records(path)
    if suffix == ".docx":
        return read_docx_records(path)
    raise ValueError(f"Формат {suffix or 'без расширения'} пока не поддерживается.")


def prepare_url_text(text: str) -> str:
    text = clean_cell(text)
    # 2GIS often stores the real URL after "?http://..." inside an internal link.
    text = re.sub(r"([?&])(https?://)", r" \2", text, flags=re.I)
    return text


def strip_url_tail(url: str) -> str:
    url = url.strip().strip("\"'<>")
    while url and url[-1] in ".,)]}!?'\"":
        url = url[:-1]
    return url


def normalize_url(raw: str, drop_query: bool = False) -> str:
    url = strip_url_tail(raw)
    if not url:
        return ""
    if url.startswith("//"):
        url = "https:" + url
    if url.startswith("www."):
        url = "https://" + url
    if not url.startswith(("http://", "https://")):
        url = "https://" + url

    parts = urlsplit(url)
    if not parts.netloc:
        return ""

    path = parts.path or ""
    query = "" if drop_query else parts.query
    normalized = urlunsplit(
        (parts.scheme.lower(), parts.netloc.lower(), path, query, "")
    )
    if not query and normalized.endswith("/") and path == "/":
        normalized = normalized[:-1]
    return normalized


def extract_urls(text: str, include_domains: bool = True) -> list[str]:
    prepared = prepare_url_text(text)
    found = [normalize_url(match.group(0)) for match in URL_RE.finditer(prepared)]
    if include_domains:
        for match in DOMAIN_RE.finditer(prepared):
            found.append(normalize_url(match.group(1)))
    return unique([url for url in found if url])


def host_of(url: str) -> str:
    try:
        return urlsplit(normalize_url(url)).netloc.lower().removeprefix("www.")
    except Exception:
        return ""


def host_matches(host: str, suffix: str) -> bool:
    return host == suffix or host.endswith("." + suffix)


def social_label(url: str) -> str:
    host = host_of(url)
    for suffix, label in SOCIAL_HOSTS.items():
        if host_matches(host, suffix):
            return label
    return ""


def is_noise_site(url: str) -> bool:
    host = host_of(url)
    if not host:
        return True
    if social_label(url):
        return True
    return any(host_matches(host, suffix) for suffix in NOISE_SITE_HOSTS)


def simplify_social_url(url: str) -> str:
    normalized = normalize_url(url, drop_query=True)
    parts = urlsplit(normalized)
    if parts.scheme == "http":
        normalized = urlunsplit(("https", parts.netloc, parts.path, parts.query, ""))
    return normalized


def find_initial_site(record: Record) -> tuple[str, str]:
    own_candidates = extract_urls(record.raw_website)
    hidden_candidates = extract_urls(record.raw_socials)

    for url in own_candidates:
        if not is_noise_site(url):
            return normalize_url(url, drop_query=True), "колонка Сайт"

    for url in hidden_candidates:
        if not is_noise_site(url):
            return normalize_url(url, drop_query=True), "из поля Соцсети 2ГИС"

    return "", ""


def extract_socials_from_text(*texts: str) -> list[str]:
    socials: list[str] = []
    for text in texts:
        for url in extract_urls(text):
            label = social_label(url)
            if label:
                socials.append(f"{label}: {simplify_social_url(url)}")
    return unique(socials)


def decode_cloudflare_email(encoded: str) -> str:
    try:
        data = bytes.fromhex(encoded)
        key = data[0]
        return "".join(chr(char ^ key) for char in data[1:])
    except Exception:
        return ""


def extract_emails(*texts: str) -> list[str]:
    emails: list[str] = []
    for text in texts:
        if not text:
            continue
        decoded = html.unescape(unquote(str(text)))
        decoded = re.sub(r"\s*(?:\[at\]|\(at\)| at )\s*", "@", decoded, flags=re.I)
        decoded = re.sub(r"\s*(?:\[dot\]|\(dot\)| dot )\s*", ".", decoded, flags=re.I)
        for cf_email in re.findall(r"data-cfemail=[\"']([0-9a-fA-F]+)[\"']", decoded):
            decoded_email = decode_cloudflare_email(cf_email)
            if decoded_email:
                emails.append(decoded_email)
        for email in EMAIL_RE.findall(decoded):
            low = email.lower().strip(".,;:()[]{}<>")
            if len(low) > 80:
                continue
            if any(noise in low for noise in EMAIL_NOISE):
                continue
            emails.append(low)
    return unique(emails)


def remove_technical_emails(emails: list[str]) -> list[str]:
    clean: list[str] = []
    for email in emails:
        domain = email.rsplit("@", 1)[-1].lower() if "@" in email else ""
        if domain in TECHNICAL_EMAIL_DOMAINS:
            continue
        clean.append(email)
    return unique(clean)


def clean_phones(*texts: str) -> list[str]:
    phones: list[str] = []
    for text in texts:
        text = clean_cell(text)
        if not text:
            continue
        text = unquote(text)
        text = re.sub(r"(?i)^tel:", "", text).strip()
        parts = re.split(r"[;\n]+", text)
        for part in parts:
            part = re.sub(r"\s+", " ", part).strip(" ,")
            if part:
                phones.append(part)
    return unique(phones)


def is_parking_page(html_text: str, final_url: str) -> bool:
    blob = html_text.lower()
    url = final_url.lower()
    return any(part in url for part in PARKING_HOST_PARTS) or any(
        word in blob for word in PARKING_WORDS
    )


def is_protected_page(html_text: str, status: int | None) -> bool:
    blob = html_text.lower()
    if status in {401, 403, 429, 451, 503}:
        return True
    return any(word in blob for word in PROTECTION_WORDS)


def classify_page(status: int | None, html_text: str, body_text: str, final_url: str) -> str:
    if is_parking_page(html_text + "\n" + body_text, final_url):
        return "паркинг/домен продается"
    if status in {404, 410}:
        return "не найден"
    if len(body_text.strip()) >= 160:
        return "живой"
    if status and 200 <= status < 400 and html_text:
        return "открылся частично"
    if is_protected_page(html_text + "\n" + body_text, status):
        return "доступ ограничен/защита"
    if status and status >= 500:
        return "ошибка сервера"
    if status is None:
        return "не открылся"
    return "непонятно"


def modernity_score(html_text: str, final_url: str) -> tuple[int | str, int | str]:
    score = 0
    html_lower = (html_text or "").lower()
    if (final_url or "").startswith("https://"):
        score += 20
    if "viewport" in html_lower:
        score += 25
    if "og:" in html_lower or 'property="og' in html_lower:
        score += 15
    if len(html_lower) > 4000:
        score += 15
    last_year = max((int(year) for year in re.findall(r"(20[12]\d)", html_lower)), default=0)
    current_year = datetime.now().year
    if last_year >= current_year - 1:
        score += 25
    elif last_year >= current_year - 3:
        score += 10
    return min(score, 100), (last_year or "")


def status_rank(status: str) -> int:
    ranks = {
        "живой": 5,
        "открылся частично": 4,
        "доступ ограничен/защита": 3,
        "ошибка сервера": 2,
        "непонятно": 1,
        "не найден": 0,
        "не открылся": 0,
        "паркинг/домен продается": 0,
    }
    return ranks.get(status, 0)


def build_url_variants(site: str) -> list[str]:
    url = normalize_url(site)
    if not url:
        return []
    parts = urlsplit(url)
    variants = [url]
    if parts.scheme == "https":
        variants.append(urlunsplit(("http", parts.netloc, parts.path, parts.query, "")))
    elif parts.scheme == "http":
        variants.append(urlunsplit(("https", parts.netloc, parts.path, parts.query, "")))
    return unique(variants)


async def collect_page(page) -> dict[str, Any]:
    try:
        html_text = await page.content()
    except Exception:
        html_text = ""
    try:
        body_text = await page.inner_text("body")
    except Exception:
        body_text = ""
    try:
        links = await page.eval_on_selector_all(
            "a[href]",
            "els => els.map(e => ({href: e.href || '', text: (e.innerText || e.textContent || '').trim()}))",
        )
    except Exception:
        links = []
    return {"html": html_text, "text": body_text, "links": links}


async def visit(context, url: str, timeout_ms: int, wait_ms: int) -> dict[str, Any]:
    page = await context.new_page()
    status = None
    error = ""
    try:
        response = await page.goto(url, wait_until="networkidle", timeout=timeout_ms)
        status = response.status if response else None
        if wait_ms:
            await page.wait_for_timeout(wait_ms)
    except Exception as exc:
        error = str(exc).splitlines()[0][:180]
        try:
            response = await page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)
            status = response.status if response else None
            if wait_ms:
                await page.wait_for_timeout(wait_ms)
            error = ""
        except Exception as fallback_exc:
            error = str(fallback_exc).splitlines()[0][:180]

    page_data = await collect_page(page)
    final_url = page.url or url
    await page.close()

    page_status = classify_page(
        status,
        page_data["html"],
        page_data["text"],
        final_url,
    )
    return {
        "url": url,
        "final_url": final_url,
        "http_status": status,
        "status": page_status,
        "html": page_data["html"],
        "text": page_data["text"],
        "links": page_data["links"],
        "error": error,
    }


def same_host(url_a: str, url_b: str) -> bool:
    return host_of(url_a) == host_of(url_b)


def contact_urls(base_url: str, links: list[dict[str, str]], limit: int) -> list[str]:
    urls: list[str] = []
    for item in links:
        href = item.get("href") or ""
        text = item.get("text") or ""
        blob = f"{href} {text}".lower()
        if not any(word in blob for word in CONTACT_WORDS):
            continue
        if href.startswith("mailto:") or href.startswith("tel:"):
            continue
        normalized = normalize_url(href, drop_query=True)
        if normalized and same_host(base_url, normalized):
            urls.append(normalized)

    for path in CONTACT_PATHS:
        urls.append(normalize_url(urljoin(base_url, path), drop_query=True))

    return unique(urls)[:limit]


def href_values(links: list[dict[str, str]]) -> list[str]:
    return [item.get("href", "") for item in links if item.get("href")]


def collect_from_page_data(data: dict[str, Any]) -> tuple[list[str], list[str], list[str]]:
    links = data.get("links") or []
    hrefs = href_values(links)
    emails = extract_emails(data.get("html", ""), data.get("text", ""), " ".join(hrefs))
    socials = extract_socials_from_text(" ".join(hrefs))
    phones = clean_phones(*(href for href in hrefs if href.lower().startswith("tel:")))
    return emails, socials, phones


async def inspect_site(browser, site: str, args) -> dict[str, Any]:
    context = await browser.new_context(
        ignore_https_errors=True,
        user_agent=USER_AGENT,
        locale="ru-RU",
        timezone_id="Europe/Moscow",
        viewport={"width": 1366, "height": 768},
    )

    async def block_heavy(route):
        if route.request.resource_type in {"image", "media", "font"}:
            await route.abort()
        else:
            await route.continue_()

    await context.route("**/*", block_heavy)

    best: dict[str, Any] | None = None
    for variant in build_url_variants(site):
        data = await visit(context, variant, args.timeout, args.wait)
        if best is None or status_rank(data["status"]) > status_rank(best["status"]):
            best = data
        if data["status"] == "живой":
            break

    if best is None:
        await context.close()
        return {
            "site": site,
            "emails": [],
            "socials": [],
            "phones": [],
            "status": "не открылся",
            "http_status": "",
            "freshness_score": "",
            "footer_year": "",
            "comment": "нет удачной попытки открыть сайт",
        }

    emails, socials, phones = collect_from_page_data(best)
    emails = remove_technical_emails(emails)

    if best["status"] in {"живой", "открылся частично"} and args.max_contact_pages > 0:
        for url in contact_urls(best["final_url"], best.get("links", []), args.max_contact_pages):
            if emails and socials:
                break
            if not same_host(best["final_url"], url):
                continue
            data = await visit(context, url, args.timeout, args.wait)
            e2, s2, p2 = collect_from_page_data(data)
            emails = remove_technical_emails(unique(emails + e2))
            socials = unique(socials + s2)
            phones = unique(phones + p2)

    await context.close()

    final_site = normalize_url(best.get("final_url") or site, drop_query=True)
    freshness_score, footer_year = ("", "")
    if best["status"] in {"живой", "открылся частично"}:
        freshness_score, footer_year = modernity_score(best.get("html", ""), final_site or site)
    comment = best.get("error", "")
    if best["status"] == "доступ ограничен/защита":
        comment = "сайт открывается с защитой; это не значит, что он мертвый"
    elif best["status"] == "паркинг/домен продается":
        comment = "похоже на припаркованный или продающийся домен"

    return {
        "site": final_site or site,
        "emails": emails,
        "socials": socials,
        "phones": phones,
        "status": best["status"],
        "http_status": best["http_status"] if best["http_status"] is not None else "",
        "freshness_score": freshness_score,
        "footer_year": footer_year,
        "comment": comment,
    }


def segment(status: str, site: str, emails: list[str], phones: list[str], socials: list[str]) -> str:
    if site and status in GOOD_SITE_STATUSES and emails:
        return "A сайт + email"
    if site and status in GOOD_SITE_STATUSES:
        return "B сайт есть, email дожать"
    if phones or socials:
        return "C без сайта, есть контакты"
    return "Брак/нет контактов"


async def process_record(index: int, total: int, record: Record, browser, sem, args) -> tuple[int, dict[str, Any]]:
    async with sem:
        site, source = find_initial_site(record)
        emails = remove_technical_emails(extract_emails(record.raw_email))
        socials = extract_socials_from_text(record.raw_socials)
        phones = clean_phones(record.raw_phone)
        status = "сайт не найден"
        http_status: int | str = ""
        freshness_score: int | str = ""
        footer_year: int | str = ""
        comment = ""

        if site:
            details = await inspect_site(browser, site, args)
            site = details["site"] or site
            emails = remove_technical_emails(unique(emails + details["emails"]))
            socials = unique(socials + details["socials"])
            if not phones:
                phones = details["phones"]
            status = details["status"]
            http_status = details["http_status"]
            freshness_score = details["freshness_score"]
            footer_year = details["footer_year"]
            comment = details["comment"]

        row = {
            "Название": record.name,
            "Сайт": site,
            "Соцсети": "; ".join(socials),
            "Ссылка 2ГИС": record.gis_link,
            "Почта электронная": "; ".join(emails),
            "Номер": "; ".join(phones),
            "Статус сайта": status,
            "HTTP код": http_status,
            "Свежесть балл": freshness_score,
            "Год в футере": footer_year,
            "Сегмент": segment(status, site, emails, phones, socials),
            "Источник сайта": source,
            "Проверено": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "Комментарий": comment,
            "Рубрики": record.categories,
            "Адрес": record.address,
            "Рейтинг": record.rating,
            "Отзывы": record.reviews,
        }

        print(
            f"[{index + 1}/{total}] {record.name[:55]} -> {status}, email: {'да' if emails else 'нет'}",
            flush=True,
        )
        return index, row


async def enrich_records(records: list[Record], args) -> list[dict[str, Any]]:
    if not records:
        return []

    try:
        from playwright.async_api import async_playwright
    except Exception as exc:
        raise RuntimeError(
            "Не установлен Playwright. Запусти run.command или выполни: "
            "python3 -m pip install -r requirements.txt && python3 -m playwright install chromium"
        ) from exc

    results: list[dict[str, Any] | None] = [None] * len(records)
    async with async_playwright() as playwright:
        browser = await playwright.chromium.launch(headless=not args.show_browser)
        sem = asyncio.Semaphore(args.workers)
        tasks = [
            process_record(index, len(records), record, browser, sem, args)
            for index, record in enumerate(records)
        ]
        for task in asyncio.as_completed(tasks):
            index, row = await task
            results[index] = row
        await browser.close()

    return [row for row in results if row is not None]


def safe_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", " ", name).strip()
    return (cleaned or "Лист")[:31]


def write_sheet(wb: Workbook, name: str, rows: list[dict[str, Any]], columns: list[str]) -> None:
    ws = wb.create_sheet(safe_sheet_name(name))
    ws.append(columns)
    for row in rows:
        ws.append([row.get(column, "") for column in columns])

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        "Название": 34,
        "Сайт": 32,
        "Соцсети": 44,
        "Ссылка 2ГИС": 38,
        "Почта электронная": 32,
        "Номер": 22,
        "Статус сайта": 24,
        "HTTP код": 10,
        "Свежесть балл": 14,
        "Год в футере": 14,
        "Сегмент": 28,
        "Источник сайта": 22,
        "Проверено": 18,
        "Комментарий": 42,
        "Рубрики": 42,
        "Адрес": 34,
        "Рейтинг": 12,
        "Отзывы": 12,
    }
    for idx, column in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(column, 20)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def save_output(rows: list[dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    extra_columns = ["Рубрики", "Адрес", "Рейтинг", "Отзывы"]
    columns = OUTPUT_COLUMNS + [column for column in extra_columns if any(row.get(column) for row in rows)]

    write_sheet(wb, "Все", rows, columns)
    for seg_name in ("A сайт + email", "B сайт есть, email дожать", "C без сайта, есть контакты", "Брак/нет контактов"):
        subset = [row for row in rows if row.get("Сегмент") == seg_name]
        if subset:
            write_sheet(wb, seg_name, subset, columns)

    wb.save(output_path)


def find_input_files(inputs: list[str]) -> list[Path]:
    files: list[Path] = []
    if inputs:
        for item in inputs:
            path = Path(item).expanduser()
            if path.is_dir():
                files.extend(
                    sorted(
                        child
                        for child in path.iterdir()
                        if child.suffix.lower() in SUPPORTED_EXTENSIONS
                    )
                )
            else:
                files.append(path)
    else:
        INPUT_DIR.mkdir(exist_ok=True)
        files = sorted(
            child
            for child in INPUT_DIR.iterdir()
            if child.suffix.lower() in SUPPORTED_EXTENSIONS
        )

    clean_files = [
        path
        for path in files
        if path.suffix.lower() in SUPPORTED_EXTENSIONS and not path.name.startswith("~$")
    ]
    missing = [str(path) for path in clean_files if not path.exists()]
    if missing:
        raise FileNotFoundError("Не найден файл: " + ", ".join(missing))
    return clean_files


def default_output_path(input_path: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    source_type = input_path.suffix.lower().lstrip(".") or "file"
    candidate = OUTPUT_DIR / f"{input_path.stem}_{source_type}_best_{stamp}.xlsx"
    counter = 2
    while candidate.exists():
        candidate = OUTPUT_DIR / f"{input_path.stem}_{source_type}_best_{stamp}_{counter}.xlsx"
        counter += 1
    return candidate


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Очищает и обогащает выгрузку 2ГИС: сайт, соцсети, email, номер, статус и A/B/C-сегмент."
    )
    parser.add_argument("inputs", nargs="*", help="Файл XLSX/CSV/JSON/DOCX, папка с файлами или пусто для папки input")
    parser.add_argument("--out", help="Куда сохранить результат. Работает только для одного входного файла.")
    parser.add_argument("--sheet", help="Название листа, если нужен не первый лист.")
    parser.add_argument("--workers", type=int, default=3, help="Сколько сайтов проверять одновременно.")
    parser.add_argument("--timeout", type=int, default=25000, help="Таймаут открытия сайта, мс.")
    parser.add_argument("--wait", type=int, default=1500, help="Сколько ждать JS после открытия, мс.")
    parser.add_argument("--max-contact-pages", type=int, default=3, help="Сколько страниц контактов проверить.")
    parser.add_argument("--show-browser", action="store_true", help="Показывать окно браузера.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if args.workers < 1:
        args.workers = 1

    try:
        input_files = find_input_files(args.inputs)
        if not input_files:
            print(
                f"Положи .xlsx, .csv, .json или .docx файл в папку {INPUT_DIR} и запусти снова.",
                file=sys.stderr,
            )
            return 2
        if args.out and len(input_files) > 1:
            print("--out можно использовать только с одним входным файлом.", file=sys.stderr)
            return 2

        for input_path in input_files:
            output_path = Path(args.out).expanduser() if args.out else default_output_path(input_path)
            print(f"\nФайл: {input_path}")
            records = read_records(input_path, args.sheet)
            print(f"Строк к обработке: {len(records)}")
            rows = asyncio.run(enrich_records(records, args))
            save_output(rows, output_path)
            print(f"Готово: {output_path}")

        return 0
    except KeyboardInterrupt:
        print("\nОстановлено пользователем.", file=sys.stderr)
        return 130
    except Exception as exc:
        message = str(exc)
        if "Executable doesn't exist" in message or "playwright install" in message:
            message = (
                "Браузер Playwright не установлен. Запусти run.command или выполни:\n"
                "python3 -m playwright install chromium"
            )
        print(f"\nОшибка: {message}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
