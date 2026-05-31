"""Codex Map Collector FastAPI server."""
from __future__ import annotations

import asyncio
import json
import os
import shutil
import sys
import uuid
from pathlib import Path
from typing import Optional

from fastapi import FastAPI, File, HTTPException, UploadFile, WebSocket, WebSocketDisconnect
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

ROOT = Path(__file__).parent
UPLOAD_DIR = ROOT / "uploads"
OUTPUT_DIR = ROOT / "output"
SETTINGS_FILE = ROOT / "data" / "settings.json"
HA_SKILLS_DIR = Path(os.getenv("HTML_ANYTHING_SKILLS_DIR", ROOT / "templates" / "html-anything" / "skills"))
UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)
(ROOT / "data").mkdir(exist_ok=True)

app = FastAPI(title="Codex Map Collector")
app.mount("/static", StaticFiles(directory=str(ROOT / "static")), name="static")
CODEX_AGENT_ID = "codex"


def codex_agent(_: str = "") -> str:
    """Public build only allows Codex."""
    return CODEX_AGENT_ID

# ── Jobs ─────────────────────────────────────────────────────────────────────
jobs: dict[str, dict] = {}

def new_job(kind: str) -> str:
    jid = str(uuid.uuid4())[:8]
    jobs[jid] = {"id": jid, "kind": kind, "status": "running", "log": [], "result": None, "task": None, "proc": None}
    return jid

def job_log(jid: str, msg: str):
    if jid in jobs:
        jobs[jid]["log"].append(msg)

def job_done(jid: str, result=None):
    if jid in jobs:
        jobs[jid]["status"] = "done"
        jobs[jid]["result"] = result

def job_fail(jid: str, err: str):
    if jid in jobs:
        jobs[jid]["status"] = "error"
        jobs[jid]["log"].append(f"❌ {err}")

@app.post("/api/job/{jid}/cancel")
async def cancel_job(jid: str):
    if jid not in jobs:
        raise HTTPException(404)
    job = jobs[jid]
    if job["status"] not in ("running", "waiting"):
        return {"ok": False, "msg": "Job already finished"}
    
    # 1. Kill subprocess if exists
    proc = job.get("proc")
    if proc:
        try:
            proc.kill()
        except Exception:
            pass
            
    # 2. Cancel async task
    task = job.get("task")
    if task:
        task.cancel()
        
    job["status"] = "canceled"
    job_log(jid, "🛑 Задача отменена пользователем")
    return {"ok": True}

@app.post("/api/shutdown")
async def shutdown_app():
    """Штатное завершение работы сервера."""
    import signal
    # Отправляем сигнал завершения самому себе
    os.kill(os.getpid(), signal.SIGTERM)
    return {"ok": True, "msg": "Shutting down..."}

# ── Settings ──────────────────────────────────────────────────────────────────
def load_settings() -> dict:
    if SETTINGS_FILE.exists():
        try:
            return json.loads(SETTINGS_FILE.read_text())
        except Exception:
            pass
    return {"selected_agent": "codex", "selected_model": "", "map_api_keys": {}}

def save_settings(data: dict):
    SETTINGS_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2))


def map_api_key(provider: str, explicit: str = "") -> str:
    if explicit:
        return explicit.strip()
    provider = (provider or "").strip().lower()
    settings_key = (load_settings().get("map_api_keys") or {}).get(provider, "")
    if settings_key:
        return str(settings_key).strip()
    env_names = {
        "2gis": "DGIS_API_KEY",
        "yandex": "YANDEX_MAPS_API_KEY",
        "google": "GOOGLE_PLACES_API_KEY",
    }
    return os.getenv(env_names.get(provider, ""), "").strip()

# ── Root ──────────────────────────────────────────────────────────────────────
@app.get("/")
async def root():
    return FileResponse(str(ROOT / "static" / "index.html"))

# ── Settings CRUD ─────────────────────────────────────────────────────────────
@app.get("/api/settings")
async def get_settings():
    return load_settings()

class SettingsPatch(BaseModel):
    selected_agent: Optional[str] = None
    selected_model: Optional[str] = None
    map_api_keys: Optional[dict[str, str]] = None

@app.patch("/api/settings")
async def patch_settings(patch: SettingsPatch):
    s = load_settings()
    if patch.selected_agent is not None:
        s["selected_agent"] = patch.selected_agent
    if patch.selected_model is not None:
        s["selected_model"] = patch.selected_model
    if patch.map_api_keys is not None:
        keys = s.get("map_api_keys") or {}
        keys.update({k: v for k, v in patch.map_api_keys.items() if isinstance(k, str)})
        s["map_api_keys"] = keys
    save_settings(s)
    return s

# ── Agents ────────────────────────────────────────────────────────────────────
@app.get("/api/agents")
async def get_agents():
    path = shutil.which("codex")
    return {"agents": [{
        "id": CODEX_AGENT_ID,
        "label": "Codex",
        "bin": "codex",
        "available": bool(path),
    }]}

# ── Upload / Table ────────────────────────────────────────────────────────────
@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...)):
    ext = Path(file.filename).suffix.lower()
    if ext not in {".xlsx", ".xls", ".csv"}:
        raise HTTPException(400, "Поддерживаются только .xlsx, .xls, .csv")
    dest = UPLOAD_DIR / f"{uuid.uuid4().hex}{ext}"
    content = await file.read()
    dest.write_bytes(content)
    data = _parse_table(dest)
    return {"path": str(dest), "name": file.filename,
            "rows": len(data["rows"]), "columns": data["columns"], "data": data}

def _parse_table(path: Path) -> dict:
    ext = path.suffix.lower()
    if ext in {".xlsx", ".xls"}:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(c) if c is not None else "" for c in next(rows_iter, [])]
        rows = [[str(c) if c is not None else "" for c in row] for row in rows_iter]
        wb.close()
    else:
        import csv
        with open(path, encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            headers = next(reader, [])
            rows = list(reader)
    return {"columns": headers, "rows": rows}

@app.get("/api/table")
async def get_table(path: str):
    p = Path(path)
    if not p.exists():
        raise HTTPException(404, "Файл не найден")
    return _parse_table(p)

class SaveTableRequest(BaseModel):
    path: str
    columns: list[str]
    rows: list[list[str]]

@app.post("/api/table/save")
async def save_table(req: SaveTableRequest):
    """Сохраняет изменения таблицы из редактора."""
    p = Path(req.path)
    ext = p.suffix.lower()
    if ext in {".xlsx", ".xls"}:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(req.columns)
        for row in req.rows:
            ws.append(row)
        wb.save(str(p))
    else:
        import csv
        with open(p, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(req.columns)
            w.writerows(req.rows)
    return {"ok": True, "path": str(p), "rows": len(req.rows)}

# ── Job status ─────────────────────────────────────────────────────────────────
@app.get("/api/job/{jid}")
async def get_job(jid: str):
    if jid not in jobs:
        raise HTTPException(404)
    j = jobs[jid]
    # Возвращаем только сериализуемые поля — task и proc не сериализуются
    return {"id": j["id"], "kind": j["kind"], "status": j["status"],
            "log": j["log"], "result": j["result"]}

# ── WebSocket log stream ───────────────────────────────────────────────────────
@app.websocket("/ws/job/{jid}")
async def ws_job(websocket: WebSocket, jid: str):
    await websocket.accept()
    sent = 0
    try:
        while True:
            job = jobs.get(jid, {})
            log = job.get("log", [])
            if len(log) > sent:
                for msg in log[sent:]:
                    await websocket.send_json({"type": "log", "msg": msg})
                sent = len(log)
            if job.get("status") in ("done", "error"):
                await websocket.send_json({"type": "status",
                                           "status": job["status"],
                                           "result": job.get("result")})
                break
            await asyncio.sleep(0.3)
    except WebSocketDisconnect:
        pass

# ── Parser ────────────────────────────────────────────────────────────────────
class ParseRequest(BaseModel):
    provider: str = "2gis"
    city: str
    query: str
    limit: int = 200
    sleep_min: float = 2.0
    sleep_max: float = 6.0
    save_raw: bool = False
    fetch_reviews: bool = False
    api_key: str = ""
    locale: str = "ru_RU"

@app.post("/api/parse")
async def start_parse(req: ParseRequest):
    jid = new_job("parse")

    async def run():
        try:
            from modules.parser.wrapper import run_parser
            provider = (req.provider or "2gis").strip().lower()
            api_key = map_api_key(provider, req.api_key)
            job_log(jid, f"Запуск парсера: {provider} / {req.city} / {req.query}")
            out_path = OUTPUT_DIR / f"parse_{jid}.csv"
            loop = asyncio.get_event_loop()
            result = await loop.run_in_executor(None, lambda: run_parser(
                city=req.city, query=req.query, limit=req.limit,
                sleep_range=(req.sleep_min, req.sleep_max),
                save_raw=req.save_raw, output_path=str(out_path),
                log_callback=lambda m: job_log(jid, m),
                fetch_reviews=req.fetch_reviews,
                provider=provider,
                api_key=api_key,
                locale=req.locale,
            ))
            job_done(jid, {"csv_path": str(out_path), "count": result.get("count", 0), "provider": provider})
        except Exception as e:
            import traceback
            job_fail(jid, traceback.format_exc()[-500:])

    jobs[jid]["task"] = asyncio.create_task(run())
    return {"job_id": jid}

# ── Enricher ──────────────────────────────────────────────────────────────────
class EnrichRequest(BaseModel):
    input_path: str
    workers: int = 3
    timeout: int = 25000
    expand_socials: bool = True
    collect_reviews: bool = True

@app.post("/api/enrich")
async def start_enrich(req: EnrichRequest):
    jid = new_job("enrich")
    out_path = OUTPUT_DIR / f"enrich_{jid}.xlsx"

    async def run():
        proc = None
        try:
            job_log(jid, f"⚡ Запускаю обогащение: {Path(req.input_path).name}")
            proc = await asyncio.create_subprocess_exec(
                sys.executable,
                str(ROOT / "modules" / "enricher" / "contact_enricher.py"),
                req.input_path, "--out", str(out_path),
                "--workers", str(req.workers), "--timeout", str(req.timeout),
                stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.STDOUT,
            )
            jobs[jid]["proc"] = proc

            # Читаем stdout с таймаутом — не висим бесконечно
            MAX_ENRICH_SECONDS = 600  # 10 минут максимум
            async def _read_stdout():
                async for line in proc.stdout:
                    job_log(jid, line.decode("utf-8", errors="replace").rstrip())

            try:
                await asyncio.wait_for(
                    asyncio.gather(_read_stdout(), proc.wait()),
                    timeout=MAX_ENRICH_SECONDS
                )
            except asyncio.TimeoutError:
                proc.kill()
                job_log(jid, f"⏱ Обогащение прервано: превышен лимит {MAX_ENRICH_SECONDS//60} мин")
                # Если файл частично создан — используем его
                if not out_path.exists():
                    job_fail(jid, "Таймаут и файл не создан")
                    return

            if proc.returncode == 0 and out_path.exists():
                job_log(jid, "✅ Проверка сайтов завершена")
                if req.expand_socials:
                    job_log(jid, "📊 Разбиваю соцсети по столбцам...")
                    _expand_socials_in_xlsx(out_path)
                job_done(jid, {"xlsx_path": str(out_path)})
                if req.collect_reviews:
                    job_log(jid, "💬 Таблица готова! Сбор отзывов из интернета продолжается в фоне...")
                    asyncio.create_task(_collect_reviews_background(out_path, jid))
            elif out_path.exists():
                # Файл есть, но код выхода не 0 — частичный результат
                job_log(jid, f"⚠️ Код выхода {proc.returncode}, но файл создан — используем")
                if req.expand_socials:
                    _expand_socials_in_xlsx(out_path)
                job_done(jid, {"xlsx_path": str(out_path)})
            else:
                job_fail(jid, f"Обогащение завершилось с кодом {proc.returncode}. Проверь лог выше.")
        except Exception as e:
            import traceback
            err = traceback.format_exc()[-600:]
            job_log(jid, f"❌ Исключение: {err}")
            if proc:
                try: proc.kill()
                except: pass
            job_fail(jid, str(e)[:200])

    jobs[jid]["task"] = asyncio.create_task(run())
    return {"job_id": jid}

async def _collect_reviews_background(path: Path, parent_jid: str):
    """Фоновый сбор отзывов — не блокирует основной джоб."""
    try:
        await _append_web_reviews_to_xlsx(path, lambda m: job_log(parent_jid, m))
        job_log(parent_jid, "✅ Фон: отзывы добавлены. Перезагрузите таблицу чтобы увидеть.")
    except Exception as e:
        job_log(parent_jid, f"⚠️ Фон: ошибка сбора отзывов: {e}")


def _expand_socials_in_xlsx(path: Path):
    """Добавляет отдельные столбцы для каждой соцсети в xlsx."""
    try:
        import openpyxl
        from modules.enricher.social_parser import parse_socials, SOCIAL_KEYS
        wb = openpyxl.load_workbook(str(path))
        for ws in wb.worksheets:
            headers = [str(c.value or "") for c in ws[1]]
            if "Соцсети" not in headers:
                continue
            soc_idx = headers.index("Соцсети")
            # Добавляем заголовки для соцсетей
            start_col = len(headers) + 1
            for i, key in enumerate(SOCIAL_KEYS):
                ws.cell(1, start_col + i, f"Соц_{key}")
            # Заполняем данные
            for row in ws.iter_rows(min_row=2):
                raw = str(row[soc_idx].value or "")
                parsed = parse_socials(raw)
                for i, key in enumerate(SOCIAL_KEYS):
                    row[0].parent.cell(row[0].row, start_col + i, parsed.get(key, ""))
        wb.save(str(path))
    except Exception as e:
        print(f"expand_socials error: {e}")

async def _append_web_reviews_to_xlsx(path: Path, log):
    """Добавляет в XLSX столбцы с отзывами/описаниями из веб-поиска."""
    import openpyxl
    from modules.enricher.web_enricher import batch_web_enrich

    wb = openpyxl.load_workbook(str(path))
    ws = wb.active
    headers = [str(c.value or "") for c in ws[1]]
    if not headers:
        return

    def col(name: str) -> int:
        nonlocal headers
        if name not in headers:
            headers.append(name)
            ws.cell(1, len(headers), name)
        return headers.index(name) + 1

    name_col = headers.index("Название") + 1 if "Название" in headers else 1
    rows = []
    row_numbers = []
    for row_num in range(2, ws.max_row + 1):
        company = {headers[i]: ws.cell(row_num, i + 1).value or "" for i in range(len(headers))}
        if str(company.get("Название", "")).strip():
            rows.append(company)
            row_numbers.append(row_num)

    if not rows:
        wb.close()
        return

    # Берём ограниченно, чтобы обычное обогащение не превращалось в многочасовой поиск.
    max_rows = min(len(rows), 80)
    log(f"  💬 Проверяю отзывы: {max_rows} строк")
    results = await batch_web_enrich(rows[:max_rows], log=log, concurrency=2)

    desc_col = col("web_description")
    reviews_json_col = col("web_reviews")
    reviews_text_col = col("Отзывы текст")
    extra_col = col("web_extra")
    sources_col = col("research_sources")
    images_col = col("research_images")
    social_col = col("research_social_text")
    md_col = col("research_md")

    for row_num, result in zip(row_numbers[:max_rows], results):
        reviews = result.get("web_reviews") or []
        review_lines = []
        for review in reviews[:6]:
            if isinstance(review, dict):
                text = str(review.get("text", "")).strip()
                author = str(review.get("author", "Клиент")).strip() or "Клиент"
                if text:
                    review_lines.append(f"{author}: {text}")
            elif review:
                review_lines.append(str(review))

        ws.cell(row_num, desc_col, result.get("web_description", ""))
        ws.cell(row_num, reviews_json_col, json.dumps(reviews, ensure_ascii=False))
        ws.cell(row_num, reviews_text_col, "\n".join(review_lines))
        ws.cell(row_num, extra_col, result.get("web_extra", ""))
        ws.cell(row_num, sources_col, json.dumps(result.get("research_sources", []), ensure_ascii=False))
        ws.cell(row_num, images_col, json.dumps(result.get("research_images", []), ensure_ascii=False))
        ws.cell(row_num, social_col, result.get("research_social_text", ""))
        ws.cell(row_num, md_col, result.get("research_md", ""))

    wb.save(str(path))
    wb.close()

# ── Web Enricher ──────────────────────────────────────────────────────────────
class WebEnrichRequest(BaseModel):
    companies: list[dict]
    concurrency: int = 3

@app.post("/api/enrich/web")
async def web_enrich(req: WebEnrichRequest):
    jid = new_job("web_enrich")

    async def run():
        try:
            from modules.enricher.web_enricher import batch_web_enrich
            job_log(jid, f"Веб-поиск для {len(req.companies)} компаний...")
            results = await batch_web_enrich(
                req.companies,
                log=lambda m: job_log(jid, m),
                concurrency=req.concurrency,
            )
            out_path = OUTPUT_DIR / f"web_enrich_{jid}.json"
            out_path.write_text(json.dumps(results, ensure_ascii=False, indent=2))
            job_log(jid, f"✅ Готово: {len(results)} компаний обогащено")
            job_done(jid, {"results": results, "json_path": str(out_path)})
        except Exception as e:
            import traceback
            job_fail(jid, traceback.format_exc()[-500:])

    jobs[jid]["task"] = asyncio.create_task(run())
    return {"job_id": jid}

# ── Landing ───────────────────────────────────────────────────────────────────
@app.get("/api/landing/styles")
async def get_landing_styles():
    from modules.landing.templates import STYLES, list_custom_templates, list_ha_templates
    styles = [{"id": k, "name": v["name"], "description": v["description"],
               "preview_color": v["preview_color"]}
              for k, v in STYLES.items()]
    custom = list_custom_templates()
    ha = list_ha_templates()
    return {"styles": styles, "custom_templates": custom, "ha_templates": ha}

class LandingRequest(BaseModel):
    company: dict
    style: str = "auto"
    agent_bin: str = "codex"
    agent_model: str = ""
    custom_template_id: str = ""
    md_dossier: str = ""
    tmpl_type: str = "standard"

class DossierRequest(BaseModel):
    company: dict

@app.post("/api/landing/dossier")
async def landing_dossier(req: DossierRequest):
    from modules.landing.generator import _build_dossier
    return {"md": _build_dossier(req.company)}

class MarketingMdRequest(BaseModel):
    companies: list[dict]
    agent_bin: str = "codex"
    agent_model: str = ""
    custom_template_id: str = ""
    custom_prompt: str = ""

@app.post("/api/landing/md")
async def landing_marketing_md(req: MarketingMdRequest):
    jid = new_job("landing_md")

    async def run():
        try:
            from modules.landing.generator import build_marketing_md
            md = await build_marketing_md(
                companies=req.companies,
                agent_bin=codex_agent(req.agent_bin),
                agent_model=req.agent_model,
                custom_template_id=req.custom_template_id,
                custom_prompt=req.custom_prompt,
                log=lambda m: job_log(jid, m),
            )
            out_path = OUTPUT_DIR / f"landing_md_{jid}.md"
            out_path.write_text(md, encoding="utf-8")
            job_done(jid, {"md": md, "md_path": str(out_path), "count": len(req.companies)})
        except Exception as e:
            import traceback
            job_fail(jid, traceback.format_exc()[-800:])

    jobs[jid]["task"] = asyncio.create_task(run())
    return {"job_id": jid}

@app.post("/api/landing/generate")
async def generate_landing(req: LandingRequest):
    jid = new_job("landing")

    async def run():
        try:
            from modules.landing.generator import generate
            result = await generate(
                company=req.company, style_id=req.style,
                agent_bin=codex_agent(req.agent_bin), agent_model=req.agent_model,
                log=lambda m: job_log(jid, m),
                output_dir=OUTPUT_DIR, job_id=jid,
                custom_template_id=req.custom_template_id,
                md_dossier=req.md_dossier,
                tmpl_type=req.tmpl_type,
            )
            job_done(jid, result)
        except Exception as e:
            import traceback
            job_fail(jid, traceback.format_exc()[-800:])

    jobs[jid]["task"] = asyncio.create_task(run())
    return {"job_id": jid}

class BatchLandingRequest(BaseModel):
    companies: list[dict]
    style: str = "auto"
    agent_bin: str = "codex"
    agent_model: str = ""
    custom_template_id: str = ""
    template_pool: list[str] = []
    concurrency: int = 2

@app.post("/api/landing/batch")
async def batch_landing(req: BatchLandingRequest):
    import random as _random
    jid = new_job("landing_batch")
    pool = req.template_pool or ([req.custom_template_id] if req.custom_template_id else [])

    async def run():
        try:
            from modules.landing.generator import generate
            sem = asyncio.Semaphore(req.concurrency)
            results = []

            async def one(company, idx):
                async with sem:
                    tmpl = _random.choice(pool) if pool else req.custom_template_id
                    name = company.get('Название', '?')
                    job_log(jid, f"[{idx+1}/{len(req.companies)}] {name}" +
                            (f" → шаблон {tmpl}" if tmpl else ""))
                    try:
                        r = await generate(
                            company=company, style_id=req.style,
                            agent_bin=codex_agent(req.agent_bin), agent_model=req.agent_model,
                            log=lambda m: job_log(jid, f"  {m}"),
                            output_dir=OUTPUT_DIR / f"batch_{jid}",
                            job_id=f"{idx}",
                            custom_template_id=tmpl,
                        )
                        results.append({"company": name, "template_used": tmpl, **r})
                    except Exception as e:
                        results.append({"company": name, "template_used": tmpl, "error": str(e)})

            await asyncio.gather(*[one(c, i) for i, c in enumerate(req.companies)])
            job_log(jid, f"✅ Пакет готов: {len(results)} лендингов")
            job_done(jid, {"results": results, "count": len(results)})
        except Exception as e:
            import traceback
            job_fail(jid, traceback.format_exc()[-500:])

    jobs[jid]["task"] = asyncio.create_task(run())
    return {"job_id": jid}

@app.get("/api/landing/preview/{jid}")
async def landing_preview(jid: str):
    path = OUTPUT_DIR / f"landing_{jid}.html"
    if not path.exists():
        raise HTTPException(404)
    return FileResponse(str(path), media_type="text/html")

@app.get("/api/landing/download/{jid}")
async def landing_download(jid: str):
    path = OUTPUT_DIR / f"landing_{jid}.html"
    if not path.exists():
        raise HTTPException(404)
    return FileResponse(str(path), media_type="text/html",
                        headers={"Content-Disposition": f'attachment; filename="landing_{jid}.html"'})

# ── Custom Templates ───────────────────────────────────────────────────────────
class TemplateRequest(BaseModel):
    name: str
    body: str

@app.post("/api/templates")
async def save_template(req: TemplateRequest):
    from modules.landing.templates import save_custom_template
    tid = save_custom_template(req.name, req.body)
    return {"id": tid, "name": req.name}

@app.get("/api/templates/hint")
async def template_hint():
    """Возвращает подсказку-промт для создания кастомного шаблона."""
    return {"hint": TEMPLATE_HINT}

TEMPLATE_HINT = """КАК СОЗДАТЬ СВОЙ ШАБЛОН ДЛЯ ЛЕНДИНГА
═══════════════════════════════════════

Шаблон — это инструкция для AI как оформить лендинг компании.
Напишите в текстовом поле что-то вроде:

───────────────────────────────────────
ПРИМЕР ШАБЛОНА: "Премиум агентство"
───────────────────────────────────────
【СТИЛЬ】 Тёмный фон (#0a0a0a), золотые акценты (#d4a017), serif-шрифты.
【НАСТРОЕНИЕ】 Роскошь, эксклюзивность, доверие состоятельных клиентов.

【СЕКЦИИ】
1. Навбар — минималистичный, тёмный, золотой логотип
2. Hero — fullscreen, крупная золотая цитата на тёмном фоне + кнопка «Записаться»
3. Наши ценности — 3 колонки с иконками, строгая типографика
4. Кейсы/портфолио — 2-3 карточки с реальными примерами (если есть данные)
5. Команда — фото-заглушки + имена (если есть)
6. Форма заявки — тёмная, минималистичная, с чекбоксом согласия
7. Футер — строгий, со всеми контактами
───────────────────────────────────────

СОВЕТЫ:
• Укажите цвета в hex (#xxxxxx) — AI будет их использовать
• Опишите настроение и целевую аудиторию
• Перечислите обязательные секции
• Можно указать шрифты: serif (Times), sans-serif (Arial), modern (Inter)
• Форма заявки всегда будет с чекбоксом согласия (по 152-ФЗ)"""

# ── Legal ──────────────────────────────────────────────────────────────────────
class LegalRequest(BaseModel):
    urls: list[str]
    agent_bin: str = "codex"
    use_ai: bool = False

@app.post("/api/legal/check")
async def legal_check(req: LegalRequest):
    # Фильтруем только живые URLs
    urls = [u for u in req.urls if u and u.startswith("http")]
    if not urls:
        raise HTTPException(400, "Нет валидных URL")
    jid = new_job("legal")

    async def run():
        try:
            from modules.legal.checker import check_sites
            job_log(jid, f"Проверка {len(urls)} сайтов по 152-ФЗ...")
            results = await check_sites(urls=urls,
                                        log=lambda m: job_log(jid, m),
                                        use_ai=req.use_ai,
                                        agent_bin=codex_agent(req.agent_bin))
            out_path = OUTPUT_DIR / f"legal_{jid}.json"
            out_path.write_text(json.dumps(results, ensure_ascii=False, indent=2))
            job_done(jid, {"results": results, "json_path": str(out_path)})
        except Exception as e:
            import traceback
            job_fail(jid, traceback.format_exc()[-500:])

    jobs[jid]["task"] = asyncio.create_task(run())
    return {"job_id": jid}

@app.get("/api/legal/memo")
async def get_memo():
    mp = ROOT / "data" / "memo_rkn.json"
    return json.loads(mp.read_text()) if mp.exists() else {"full_text": ""}


# ── Screenshots ────────────────────────────────────────────────────────────────
class ScreenshotRequest(BaseModel):
    urls: list[str]
    generate_concept: bool = False
    openai_api_key: str = ""
    concept_prompt: str = ""
    agent_bin: str = "codex"

@app.post("/api/screenshots/capture")
async def capture_screenshots(req: ScreenshotRequest):
    jid = new_job("screenshots")

    async def run():
        try:
            from modules.screenshots.capture import capture_and_analyze
            results = await capture_and_analyze(
                urls=req.urls, output_dir=OUTPUT_DIR / f"screens_{jid}",
                generate_concept=req.generate_concept,
                openai_api_key=req.openai_api_key,
                concept_prompt=req.concept_prompt,
                agent_bin=codex_agent(req.agent_bin),
                log=lambda m: job_log(jid, m),
            )
            job_done(jid, results)
        except Exception as e:
            import traceback
            job_fail(jid, traceback.format_exc()[-500:])

    jobs[jid]["task"] = asyncio.create_task(run())
    return {"job_id": jid}

@app.get("/api/screenshots/{jid}/{filename}")
async def get_screenshot(jid: str, filename: str):
    path = OUTPUT_DIR / f"screens_{jid}" / filename
    if not path.exists():
        raise HTTPException(404)
    return FileResponse(str(path))

class AnalyzeSiteRequest(BaseModel):
    url: str
    company: dict = {}
    page_text: str = ""
    freshness: int = 0
    agent_bin: str = "codex"
    agent_model: str = ""
    openai_api_key: str = ""
    screenshots_job_id: str = ""

@app.post("/api/screenshots/analyze_site")
async def analyze_site(req: AnalyzeSiteRequest):
    jid = new_job("screen_analyze")
    output_dir = OUTPUT_DIR / f"screens_{req.screenshots_job_id}" if req.screenshots_job_id else OUTPUT_DIR / f"analysis_{jid}"

    async def run():
        try:
            from modules.screenshots.capture import analyze_site_with_codex
            result = await analyze_site_with_codex(
                url=req.url,
                company=req.company,
                page_text=req.page_text,
                freshness=req.freshness,
                agent_bin=codex_agent(req.agent_bin),
                agent_model=req.agent_model,
                openai_api_key=req.openai_api_key,
                output_dir=output_dir,
                log=lambda m: job_log(jid, m),
            )
            job_done(jid, result)
        except Exception as e:
            import traceback
            job_fail(jid, traceback.format_exc()[-500:])

    jobs[jid]["task"] = asyncio.create_task(run())
    return {"job_id": jid}

# ── Chat ───────────────────────────────────────────────────────────────────────
class ChatRequest(BaseModel):
    tab: str
    message: str
    context: dict = {}

TAB_CONTEXT = {
    "parser": "Вкладка ПАРСИНГ — сбор данных компаний из 2GIS, Yandex Maps, Google Places или OpenStreetMap.",
    "enrich": "Вкладка ОБОГАЩЕНИЕ — проверка сайтов, разбивка соцсетей, дообогащение из интернета, редактор таблицы.",
    "legal": "Вкладка АУДИТ 152-ФЗ — проверка сайтов компаний на нарушения закона о персональных данных.",
    "landing": "Вкладка ЛЕНДИНГИ — генерация лендингов для компаний без сайта через Codex с шаблонами.",
    "screens": "Вкладка АНАЛИЗ САЙТОВ — скриншоты устаревших сайтов и генерация нового концепта через GPT Image.",
    "send": "Вкладка ОТПРАВИТЬ — управление отправкой материалов компаниям через все найденные каналы связи.",
}

@app.post("/api/chat/{tab}")
async def chat(tab: str, req: ChatRequest):
    """Принимает сообщение пользователя и форматирует его в структурированный запрос к разработчику."""
    tab_ctx = TAB_CONTEXT.get(tab, f"Вкладка {tab}")
    ctx_str = ""
    if req.context:
        ctx_str = "\nКонтекст: " + ", ".join(f"{k}={v}" for k, v in req.context.items() if v)

    structured = _format_dev_request(req.message, tab_ctx, ctx_str)
    return {"response": structured, "tab": tab}

def _format_dev_request(msg: str, tab_ctx: str, context: str) -> str:
    """Форматирует пользовательское сообщение в структурированный запрос."""
    msg = msg.strip()
    if not msg:
        return ""

    # Определяем тип запроса
    is_bug = any(w in msg.lower() for w in ["не работает", "ошибка", "баг", "сломал", "не открывается", "не создаёт", "пустой", "не сохраняется"])
    is_feature = any(w in msg.lower() for w in ["хочу", "добавь", "нужно", "сделай", "хотелось бы", "можно ли"])
    is_question = msg.endswith("?") or any(w in msg.lower() for w in ["как", "почему", "зачем", "где"])

    req_type = "🐛 БАГ" if is_bug else ("✨ ФИЧА" if is_feature else ("❓ ВОПРОС" if is_question else "📝 ЗАПРОС"))

    return f"""{req_type} | {tab_ctx}

**Сообщение:** {msg}
{context}

**Для разработчика:**
— Контекст: Codex Map Collector, {tab_ctx}
— Тип: {req_type}
— Описание: {msg}
— Приоритет: {"Высокий (мешает работе)" if is_bug else "Средний"}

Скопируйте этот блок и отправьте разработчику."""

# ── html-anything Templates Gallery ───────────────────────────────────────────
import re as _re

def _parse_skill_meta(text: str) -> dict:
    meta: dict = {}
    m = _re.match(r'^---\s*\n(.*?)\n---', text, _re.DOTALL)
    if m:
        for line in m.group(1).splitlines():
            if ':' in line:
                k, _, v = line.partition(':')
                meta[k.strip()] = v.strip().strip('"')
    return meta

@app.get("/api/ha/templates")
async def ha_templates_list():
    """Возвращает все 75 шаблонов из html-anything с метаданными."""
    if not HA_SKILLS_DIR.exists():
        return {"templates": []}
    result = []
    for d in sorted(HA_SKILLS_DIR.iterdir()):
        if not d.is_dir(): continue
        skill_md = d / "SKILL.md"
        if not skill_md.exists(): continue
        try:
            text = skill_md.read_text(encoding="utf-8", errors="replace")
            meta = _parse_skill_meta(text)
            result.append({
                "id": d.name,
                "emoji": meta.get("emoji", "📄"),
                "name": meta.get("en_name", d.name),
                "zh_name": meta.get("zh_name", ""),
                "description": meta.get("description", ""),
                "category": meta.get("category", "other"),
                "scenario": meta.get("scenario", ""),
                "aspect_hint": meta.get("aspect_hint", ""),
                "tags": meta.get("tags", "").strip("[]").replace('"', "").split(",") if meta.get("tags") else [],
                "featured": int(meta.get("featured", 0) or 0),
                "has_preview": (d / "example.html").exists(),
            })
        except Exception:
            continue
    # Сортируем: featured сначала, потом по алфавиту
    result.sort(key=lambda x: (x["featured"] == 0, x["name"]))
    return {"templates": result, "count": len(result)}

@app.get("/api/ha/templates/{template_id}/preview")
async def ha_template_preview(template_id: str):
    """Возвращает example.html для превью шаблона."""
    # Защита от path traversal
    safe_id = _re.sub(r'[^a-zA-Z0-9\-_]', '', template_id)
    html_path = HA_SKILLS_DIR / safe_id / "example.html"
    if not html_path.exists():
        raise HTTPException(404, "Preview not found")
    return FileResponse(str(html_path), media_type="text/html")

@app.get("/api/ha/templates/{template_id}/skill")
async def ha_template_skill(template_id: str):
    """Возвращает тело SKILL.md (промт)."""
    safe_id = _re.sub(r'[^a-zA-Z0-9\-_]', '', template_id)
    skill_path = HA_SKILLS_DIR / safe_id / "SKILL.md"
    if not skill_path.exists():
        raise HTTPException(404)
    text = skill_path.read_text(encoding="utf-8")
    body = _re.sub(r'^---\s*\n.*?\n---\s*\n', '', text, flags=_re.DOTALL).strip()
    return {"id": safe_id, "body": body}

# ── Landing Refinement (chat with AI about specific landing) ───────────────────
class RefineRequest(BaseModel):
    job_id: str
    instruction: str
    agent_bin: str = "codex"
    agent_model: str = ""

@app.post("/api/landing/refine")
async def refine_landing(req: RefineRequest):
    """Редактирует существующий лендинг по инструкции пользователя."""
    html_path = OUTPUT_DIR / f"landing_{req.job_id}.html"
    if not html_path.exists():
        raise HTTPException(404, "Лендинг не найден")

    jid = new_job("landing_refine")
    current_html = html_path.read_text(encoding="utf-8")

    async def run():
        try:
            from modules.landing.generator import _call_agent
            job_log(jid, f"Редактирование лендинга по запросу: {req.instruction[:80]}...")

            prompt = f"""Ты — эксперт по веб-дизайну. Тебе дан существующий HTML лендинга.
Внеси МИНИМАЛЬНЫЕ изменения согласно инструкции пользователя.
Выведи ПОЛНЫЙ обновлённый HTML (весь документ от <!DOCTYPE до </html>).
Не объясняй изменения — только HTML.

ИНСТРУКЦИЯ ПОЛЬЗОВАТЕЛЯ: {req.instruction}

ТЕКУЩИЙ HTML:
{current_html[:20000]}
"""
            new_html = await _call_agent(
                agent_bin=codex_agent(req.agent_bin),
                agent_model=req.agent_model,
                prompt=prompt,
                log=lambda m: job_log(jid, m),
                timeout=120,
            )
            if new_html and new_html.strip().startswith("<"):
                html_path.write_text(new_html, encoding="utf-8")
                job_log(jid, "✅ Лендинг обновлён")
                job_done(jid, {"html_path": str(html_path), "updated": True})
            else:
                job_log(jid, "⚠️ AI не вернул HTML, лендинг не изменён")
                job_done(jid, {"updated": False})
        except Exception as e:
            import traceback
            job_fail(jid, traceback.format_exc()[-400:])

    jobs[jid]["task"] = asyncio.create_task(run())
    return {"job_id": jid, "landing_job_id": req.job_id}

# ── Owner Search (добавить поиск владельца при обогащении) ────────────────────
class OwnerSearchRequest(BaseModel):
    company_name: str
    address: str = ""

@app.post("/api/enrich/owner")
async def find_owner(req: OwnerSearchRequest):
    """Ищет имя владельца/руководителя компании."""
    try:
        import httpx
        query = f"{req.company_name} {req.address} руководитель директор владелец"
        async with httpx.AsyncClient(timeout=10, verify=False) as client:
            resp = await client.get(
                "https://html.duckduckgo.com/html/",
                params={"q": query, "kl": "ru-ru"},
                headers={"User-Agent": "Mozilla/5.0"}
            )
        import re
        snippets = re.findall(r'class="result__snippet"[^>]*>(.*?)</a>', resp.text, re.DOTALL)
        clean = [re.sub(r'<[^>]+>', '', s).strip() for s in snippets[:3]]
        # Ищем имена ФИО
        name_re = re.compile(r'[А-ЯЁ][а-яё]+\s+[А-ЯЁ][а-яё]+(?:\s+[А-ЯЁ][а-яё]+)?')
        names = []
        for snippet in clean:
            names.extend(name_re.findall(snippet))
        return {"names": list(set(names))[:3], "snippets": clean}
    except Exception as e:
        return {"names": [], "error": str(e)}

# ── Download ───────────────────────────────────────────────────────────────────
@app.get("/api/download")
async def download_file(path: str):
    p = Path(path)
    if not p.exists():
        raise HTTPException(404)
    return FileResponse(str(p), headers={
        "Content-Disposition": f'attachment; filename="{p.name}"'
    })
